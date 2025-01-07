import logging
import openpyxl
import re

oPrintToLogLogger = logging.getLogger("print_to_log")

class Signal():
    def __init__(self, name, value, min_, max_):
        self.name   = name
        self.min_   = min_
        self.max_   = max_

class RangeCheck():
    nl  =   "\n"
    tab =   "    "
    min_min = "-1E+10"
    max_max = "1E+10"
    def __init__(self, signal_list, config_xml):
        self.signal_list    = signal_list
        self.config_xml     = config_xml
        self.signals            = []
        self.signals_project    = []
        try:
            self.parse_signals_from_config_xml()
            self.parse_signal_list()
            self.create_check_range_python()
        except Exception as e:
            oPrintToLogLogger.warning(self.tab + "Could not create signalRangeCheck.py.")
            oPrintToLogLogger.warning(self.tab + str(e))

    @staticmethod
    def get_spaces(last_string, max_length):
        spaces = ""                          
        if max_length > len(last_string):
            spaces = " "*(max_length - len(last_string))
        return last_string + spaces
        
    def parse_signals_from_config_xml(self):
        signal_re = [r".*?<Signal name=\"(\w+)\",.*",
        r".*?<Constant name=\"(\w+)\"",
        r".*?<OpenPort name=\"(\w+)\""]
        with open(self.config_xml, "r") as config_obj:
            for line in config_obj:
                for regex in signal_re:
                    match = re.search(regex, line)
                    if match:
                        self.signals_project.append(match.group(1))
                        break
            
    def parse_signal_list(self):
        wb_obj = openpyxl.load_workbook(self.signal_list)
        sheet = wb_obj.active
        nrows = sheet.max_row
        for i in range(1, nrows):
            if sheet.cell(i+1, 1).value != None:
                name    = sheet.cell(i+1, 1).value.replace(" ", "").replace("\t", "").replace("\n", "") 
                min     = str(sheet.cell(i+1, 6).value).replace(" ", "").replace("\t", "").replace("\n", "")  if sheet.cell(i+1, 6).value != None else self.min_min
                min     = min.replace("-Inf", self.min_min)
                max     = str(sheet.cell(i+1, 7).value).replace(" ", "").replace("\t", "").replace("\n", "")  if sheet.cell(i+1, 7).value != None else self.max_max
                max     = max.replace("Inf", self.max_max)
                try:
                    init = str((float(min) + float(max))*0.5)
                except Exception as e:
                    oPrintToLogLogger.warning(self.tab + "Probably a weird entry in the Excel signal list for {name}.".format(name=name))
                    oPrintToLogLogger.warning(self.tab + str(e))
                    continue
                signal = Signal(name, init, min, max)    
                self.signals.append(signal)
            else:
                pass # ignoring None-valued cells.
        
    def create_check_range_python(self):
        python_lines    = []
        interface_lines = []
        main_lines      = []
        interface_lines.append("time = Variable(\"currentTime\")")
        for i, signal in enumerate(self.signals):
            if signal.name in self.signals_project:
                interface_lines.append(self.get_spaces(signal.name, 50) + " = Variable(\"" + signal.name + "\")")
                main_lines.append(self.tab + "if " + self.get_spaces(signal.name+".Value", 50) + 
                "<" + self.get_spaces(signal.min_, 15) + " or " + self.get_spaces(signal.name+".Value", 50) + " > " + 
                self.get_spaces(signal.max_+ ":", 15) + 
                " logThis(ERROR_ERROR, \"Range Error: " + self.get_spaces(signal.name, 50) + "is not within " + 
                self.get_spaces(signal.min_, 15) + " and " + 
                signal.max_  +"\")")
        python_lines.append("from synopsys.silver import *")
        python_lines.extend(interface_lines)
        python_lines.append("def MainGenerator(*args):")
        python_lines.extend(main_lines)
        
        with open("signalRangeCheck.py", "w") as py_file:
            py_file.write("\n".join(python_lines))
            